import os
import openpyxl
import openpyxl.styles



os.chdir(r"E:\桌面")  # 创建文件路径
workbook = openpyxl.Workbook()  # 创建excel
sheet = workbook.active


titles_1 = ['贴片1', '贴片2', '引线键合1']
titles_2 = ['生产机台', '消耗物料', '消耗物料批次', '进站数量', '出站数量']
column2 = 1
column1 = 1
# 工序标题填充
for k in range(len(titles_1)):
    tem = column1+len(titles_2) # 开始时为6
    # print(tem)
    # 从第一行第一列开始到第一行第五列进行合并操作
    sheet.merge_cells(start_row=1, start_column=column1, end_row=1, end_column=tem-1)
    # 从第一行第一列开始进行填充数据信息
    cell_a = sheet.cell(row=1, column=column1)
    cell_a.value = titles_1[k] # 根据下表获取数据
    column1 = tem # 每次循环将已经使用的单元格长度进行求和
    print(column1)
    print('*'*50)
    for i in range(len(titles_2)):
        # 从第二行第一列开始进行填充数据信息
        cell_a1 = sheet.cell(row=2, column=column2)
        # 根据下表获取数据
        cell_a1.value = titles_2[i]
        # 每次循环后将列数加一
        column2 += 1

workbook.save('2.xlsx')
